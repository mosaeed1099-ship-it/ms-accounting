"""
Import monthly fees from Excel → /api/monthly-fees/import
"""
import sys, time, requests
from openpyxl import load_workbook

BASE  = "https://ms-accounting-api-production.up.railway.app"
EMAIL = "ms.owner@mshq.io"
PASS  = "MS@QVj8ebqSw1iAOdLR#26"
XLSX  = "/Users/render/Downloads/مراجعة المستحقات الشهرية.xlsx"

SHEETS_ORDER = [
    ('أكتوبر 2025', 'شهر اكتوبر',   2025, 10),
    ('نوفمبر 2025', 'شهر نوفمبر',   2025, 11),
    ('ديسمبر 2025', 'شهر ديسمبر',   2025, 12),
    ('يناير 2026',  'يناير 2026',    2026,  1),
    ('فبراير 2026', 'فبراير  2026',  2026,  2),
    ('مارس 2026',   'مارس  2026 ',   2026,  3),
    ('أبريل 2026',  'ابريل  2026 ',  2026,  4),
    ('مايو 2026',   'مايو  2026 ',   2026,  5),
]


def login():
    r = requests.post(f"{BASE}/api/auth/login",
                      data={"username": EMAIL, "password": PASS},
                      headers={"Content-Type": "application/x-www-form-urlencoded"})
    token = r.json().get("access_token")
    if not token:
        print("❌ Login failed:", r.text); sys.exit(1)
    return {"Authorization": f"Bearer {token}"}


def row_fill_type(ws, row_idx):
    fills = []
    for col in range(1, 8):
        c = ws.cell(row=row_idx, column=col)
        if c.fill and c.fill.fgColor and c.fill.fgColor.type == 'rgb':
            rgb = c.fill.fgColor.rgb.upper()
            if 'FF0000' in rgb:   fills.append('red')
            elif 'FFFF00' in rgb: fills.append('yellow')
    red_cols = fills.count('red')
    return 'red' if red_cols >= 5 else 'normal'


def parse_excel():
    wb = load_workbook(XLSX)
    sheets_data = {}
    for label, sn, yr, mo in SHEETS_ORDER:
        ws = wb[sn]
        clients = []
        for row_idx in range(4, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, 9)]
            if row[0] is None: continue
            try: int(str(row[0]).strip())
            except: continue
            name = str(row[1]).strip() if row[1] else ''
            if not name or name == 'None': continue
            fee   = float(row[4] or 0)
            bayan = str(row[6]).strip() if row[6] else ''
            paid  = bool(bayan and 'دفع' in bayan)
            red   = row_fill_type(ws, row_idx) == 'red'
            clients.append({'name': name, 'fee': fee, 'bayan': bayan, 'paid': paid, 'red': red})
        sheets_data[(yr, mo)] = clients
    return sheets_data


def build_payload(sheets_data):
    # مايو = آخر شهر → يحدد الحالة
    mayo = sheets_data[(2026, 5)]
    red_in_mayo = {c['name'] for c in mayo if c['red']}

    # كل العملاء الفريدين
    all_clients = {}
    for (yr, mo), clients in sheets_data.items():
        for c in clients:
            if c['name'] not in all_clients:
                all_clients[c['name']] = {'name': c['name'], 'monthly_fee': c['fee'], 'red': c['red']}
            else:
                # آخر قيمة أتعاب تطغى
                all_clients[c['name']]['monthly_fee'] = c['fee']

    # تحديد الحالة: أحمر في مايو → archived
    client_payloads = []
    for name, info in all_clients.items():
        status = "archived" if name in red_in_mayo else "active"
        client_payloads.append({
            "name": name,
            "monthly_fee": info['monthly_fee'],
            "status": status,
        })

    # سجلات الدفع
    record_payloads = []
    for (yr, mo), clients in sheets_data.items():
        for c in clients:
            record_payloads.append({
                "client_name": c['name'],
                "year": yr,
                "month": mo,
                "fee_amount": c['fee'],
                "paid": c['paid'],
                "bayan": c['bayan'] or None,
            })

    return client_payloads, record_payloads


def wait_for_api(headers, retries=15):
    for i in range(retries):
        try:
            r = requests.get(f"{BASE}/api/monthly-fees/clients", headers=headers, timeout=10)
            if r.status_code in (200, 403):
                return True
        except Exception:
            pass
        print(f"  ⏳ انتظار Railway... ({i+1}/{retries})")
        time.sleep(10)
    return False


def main():
    print("🔐 تسجيل الدخول...")
    headers = login()

    print("⏳ انتظار الـ API الجديد على Railway...")
    if not wait_for_api(headers):
        print("❌ الـ API لم يستجب — ادفع مرة ثانية بعد دقيقتين")
        sys.exit(1)
    print("✅ الـ API جاهز")

    print("📊 قراءة الشيت...")
    sheets_data = parse_excel()
    clients, records = build_payload(sheets_data)

    active   = [c for c in clients if c['status'] == 'active']
    archived = [c for c in clients if c['status'] == 'archived']
    print(f"   {len(clients)} عميل فريد — نشط: {len(active)} | أرشيف: {len(archived)}")
    print(f"   {len(records)} سجل دفع")

    # Pass 1: clients only (no records) to create them fast
    print("📤 [1/2] استيراد العملاء...")
    r = requests.post(f"{BASE}/api/monthly-fees/import",
                      json={"clients": clients, "records": []},
                      headers=headers, timeout=30)
    if r.status_code != 200:
        print(f"❌ فشل استيراد العملاء: {r.status_code} {r.text[:200]}")
        sys.exit(1)
    print(f"   ✅ عملاء: {r.json()['clients_imported']}")

    # Pass 2: records in batches of 50
    BATCH = 50
    total_saved = 0
    for i in range(0, len(records), BATCH):
        batch = records[i:i+BATCH]
        r = requests.post(f"{BASE}/api/monthly-fees/import",
                          json={"clients": [], "records": batch},
                          headers=headers, timeout=60)
        if r.status_code != 200:
            print(f"❌ فشل batch {i}: {r.status_code} {r.text[:200]}")
            continue
        saved = r.json().get('records_saved', 0)
        total_saved += saved
        print(f"   ✅ [{i+1}→{min(i+BATCH,len(records))}] سجلات محفوظة: {saved}")
    print(f"   إجمالي السجلات: {total_saved}/{len(records)}")

    # تحقق سريع
    r2 = requests.get(f"{BASE}/api/monthly-fees/clients", headers=headers)
    print(f"   إجمالي العملاء في DB: {len(r2.json())}")


if __name__ == "__main__":
    main()
