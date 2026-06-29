"""
/api/vat-excel — VAT Excel analysis router.
Accepts ETA e-invoicing Excel exports, parses them with the Universal Excel Engine,
stores results per-user, and serves history/drill/declaration endpoints.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.deps import get_current_user
from app.database import get_db
from app.models.user import User

router = APIRouter(prefix="/api/vat-excel", tags=["vat_excel"])
logger = logging.getLogger("vat_excel")

# ── DB table (auto-create) ─────────────────────────────────────────────────────
_TABLE_READY = False


def _ensure_table(db: Session):
    global _TABLE_READY
    if _TABLE_READY:
        return
    stmts = [
        """CREATE TABLE IF NOT EXISTS vat_excel_analyses (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER,
            created_at   TIMESTAMP DEFAULT NOW(),
            company_name VARCHAR(300),
            tax_number   VARCHAR(80),
            period_label VARCHAR(80),
            net_vat      NUMERIC(14,2) DEFAULT 0,
            sales_net    NUMERIC(14,2) DEFAULT 0,
            sales_vat    NUMERIC(14,2) DEFAULT 0,
            pur_net      NUMERIC(14,2) DEFAULT 0,
            pur_vat      NUMERIC(14,2) DEFAULT 0,
            total_invoices INTEGER DEFAULT 0,
            data_json    TEXT
        )""",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS year INTEGER",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS month INTEGER",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS net_vat NUMERIC(14,2) DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS sales_net NUMERIC(14,2) DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS sales_vat NUMERIC(14,2) DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS pur_net NUMERIC(14,2) DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS pur_vat NUMERIC(14,2) DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS total_invoices INTEGER DEFAULT 0",
        "ALTER TABLE vat_excel_analyses ADD COLUMN IF NOT EXISTS data_json TEXT",
    ]
    for sql in stmts:
        try:
            db.execute(text(sql))
            db.commit()
        except Exception:
            db.rollback()
    _TABLE_READY = True


# ── Flexible ETA parser ────────────────────────────────────────────────────────

def _find_col(headers: list, *fragments: str) -> Optional[int]:
    """Return first column index whose header contains any of the given fragments."""
    for frag in fragments:
        frag_lower = frag.lower()
        for i, h in enumerate(headers):
            if isinstance(h, str) and frag_lower in h.lower():
                return i
    return None


def _num(val) -> float:
    try:
        return float(str(val).replace(",", "").replace("،", "").strip())
    except Exception:
        return 0.0


def _str(val) -> str:
    s = str(val).strip()
    return "" if s in ("nan", "None", "NaT") else s


def _parse_date_str(val) -> str:
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s if s not in ("nan", "None", "") else ""


def _parse_date_obj(val) -> Optional[date]:
    s = _parse_date_str(val)
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _detect_header_row(df) -> int:
    """Try excel_engine, fall back to scanning first 10 rows."""
    try:
        from app.routers.excel_engine import detect_header_row
        return detect_header_row(df)
    except Exception:
        pass
    # Fallback: first row with >= 4 non-null string cells
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        str_cells = sum(1 for v in row if isinstance(v, str) and len(v.strip()) > 1)
        if str_cells >= 4:
            return i
    return 0


def _parse_excel(content: bytes) -> dict:
    import pandas as pd

    try:
        xl = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None, dtype=str)
    except Exception as e:
        raise HTTPException(400, f"تعذر قراءة الملف: {e}")

    if not xl:
        raise HTTPException(400, "الملف فارغ")

    # 1. Find invoice sheet
    invoice_sheet = None
    for name in ["جميع الفواتير", "All Invoices", "Invoices", "فواتير"]:
        if name in xl and not xl[name].empty:
            invoice_sheet = xl[name]
            break

    if invoice_sheet is None:
        # Pick largest sheet by row count
        best = max(xl.items(), key=lambda x: len(x[1]))
        invoice_sheet = best[1]

    if invoice_sheet is None or invoice_sheet.empty:
        raise HTTPException(400, "الملف لا يحتوي على بيانات فواتير")

    # 2. Detect header row
    header_row = _detect_header_row(invoice_sheet)
    headers = [str(h) if h is not None else "" for h in invoice_sheet.iloc[header_row].tolist()]
    data = invoice_sheet.iloc[header_row + 1:].copy().reset_index(drop=True)

    # Drop fully-empty rows and totals rows
    data = data.dropna(how="all")
    for totals_kw in ["الإجمالي", "المجموع", "Total", "Grand Total"]:
        mask = data.apply(lambda row: any(totals_kw in str(v) for v in row), axis=1)
        data = data[~mask]
    data = data.reset_index(drop=True)

    if data.empty:
        raise HTTPException(400, "لم يتم العثور على صفوف فواتير بعد معالجة الملف")

    # 3. Map columns flexibly
    col = {
        "uuid":         _find_col(headers, "uuid", "UUID"),
        "internal_id":  _find_col(headers, "رقم الفاتورة الداخلي", "internal", "رقم الفاتورة"),
        "doc_type":     _find_col(headers, "نوع المستند", "document type", "doc type"),
        "direction":    _find_col(headers, "اتجاه", "direction", "نوع الفاتورة"),
        "status":       _find_col(headers, "حالة الفاتورة", "status", "الحالة"),
        "issue_date":   _find_col(headers, "تاريخ الإصدار", "issue date", "التاريخ", "date"),
        "submit_date":  _find_col(headers, "تاريخ التقديم", "submit date"),
        "issuer":       _find_col(headers, "اسم البائع", "issuer name", "المورد"),
        "issuer_tin":   _find_col(headers, "رقم ضريبي البائع", "issuer tin"),
        "receiver":     _find_col(headers, "اسم العميل", "اسم المشتري", "receiver name", "العميل"),
        "receiver_tin": _find_col(headers, "الرقم الضريبي للعميل", "receiver tin"),
        "net":          _find_col(headers, "المبلغ الصافي", "net amount", "الصافي"),
        "discount":     _find_col(headers, "الخصم", "discount"),
        "net_after":    _find_col(headers, "الإجمالي قبل الضريبة", "net after", "المبلغ بعد الخصم"),
        "vat":          _find_col(headers, "إجمالي ضريبة القيمة المضافة", "vat amount", "ضريبة القيمة", "الضريبة"),
        "total":        _find_col(headers, "الإجمالي النهائي", "total amount", "الإجمالي"),
        "currency":     _find_col(headers, "العملة", "currency"),
    }

    # 4. Parse rows
    rows = []
    for _, row in data.iterrows():
        def cv(c):
            return _str(row.iloc[c]) if c is not None and c < len(row) else ""
        def nv(c):
            return _num(row.iloc[c]) if c is not None and c < len(row) else 0.0

        net_val     = nv(col["net"])
        discount    = nv(col["discount"])
        net_after   = nv(col["net_after"]) if col["net_after"] is not None else (net_val - discount)
        vat_val     = nv(col["vat"])
        total_val   = nv(col["total"])
        direction   = cv(col["direction"])
        status_val  = cv(col["status"])
        issue_date  = cv(col["issue_date"])
        if issue_date:
            issue_date = _parse_date_str(issue_date)

        rows.append({
            "uuid":        cv(col["uuid"]),
            "internal_id": cv(col["internal_id"]),
            "doc_type":    cv(col["doc_type"]),
            "direction":   direction,
            "status":      status_val,
            "issue_date":  issue_date,
            "submit_date": _parse_date_str(cv(col["submit_date"])) if col["submit_date"] is not None else "",
            "issuer":      cv(col["issuer"]),
            "issuer_tin":  cv(col["issuer_tin"]),
            "receiver":    cv(col["receiver"]),
            "receiver_tin":cv(col["receiver_tin"]),
            "net":         net_val,
            "discount":    discount,
            "net_after":   net_after,
            "vat":         vat_val,
            "total":       total_val,
            "currency":    cv(col["currency"]) or "EGP",
        })

    if not rows:
        raise HTTPException(400, "لا توجد فواتير داخل الملف بعد التحليل")

    # 5. Detect period from dates
    valid_dates = [_parse_date_obj(r["issue_date"]) for r in rows if r["issue_date"]]
    year, month = None, None
    if valid_dates:
        most_common = Counter((d.year, d.month) for d in valid_dates if d).most_common(1)
        if most_common:
            year, month = most_common[0][0]

    MONTH_AR = {1:"يناير",2:"فبراير",3:"مارس",4:"أبريل",5:"مايو",6:"يونيو",
                7:"يوليو",8:"أغسطس",9:"سبتمبر",10:"أكتوبر",11:"نوفمبر",12:"ديسمبر"}
    period_label = f"{MONTH_AR.get(month, str(month) if month else '?')} {year}" if year else "غير محدد"

    # 6. Extract company info from other sheets
    company_name = ""
    tax_number   = ""
    for sheet_name, sheet_df in xl.items():
        if sheet_name == "جميع الفواتير":
            continue
        for r_idx in range(min(20, len(sheet_df))):
            row_vals = [_str(v) for v in sheet_df.iloc[r_idx]]
            for v in row_vals:
                if any(kw in v for kw in ["الرقم الضريبي", "tax id", "tin"]):
                    # Next cell might be the tax number
                    idx_v = row_vals.index(v)
                    if idx_v + 1 < len(row_vals) and row_vals[idx_v + 1]:
                        tax_number = row_vals[idx_v + 1]
                if any(kw in v for kw in ["اسم الشركة", "company name", "الممول"]):
                    idx_v = row_vals.index(v)
                    if idx_v + 1 < len(row_vals) and row_vals[idx_v + 1]:
                        company_name = row_vals[idx_v + 1]
        if company_name and tax_number:
            break

    # Try header area of invoice sheet for company info
    if not company_name:
        for r_idx in range(min(header_row, 10)):
            row_vals = [_str(v) for v in invoice_sheet.iloc[r_idx]]
            for v in row_vals:
                if v and len(v) > 3 and not v.startswith("nan"):
                    if any(kw in v for kw in ["اسم", "company", "ممول"]):
                        company_name = v
                    elif any(c.isdigit() for c in v) and len(v) >= 9:
                        tax_number = v

    # 7. Aggregate
    valid_status = {"صالحة", "معتمدة", "مقبولة", "Valid", "Accepted", ""}
    def is_valid(r):
        return r["status"] in valid_status or not r["status"]

    outgoing = [r for r in rows if "صادرة" in r["direction"] and is_valid(r)]
    incoming = [r for r in rows if "واردة" in r["direction"] and is_valid(r)]

    sales_net = round(sum(r["net_after"] or r["net"] for r in outgoing), 2)
    sales_vat = round(sum(r["vat"] for r in outgoing), 2)
    pur_net   = round(sum(r["net_after"] or r["net"] for r in incoming), 2)
    pur_vat   = round(sum(r["vat"] for r in incoming), 2)
    net_vat   = round(sales_vat - pur_vat, 2)

    all_amounts = [r["total"] for r in rows if r["total"] > 0]
    max_invoice = max(all_amounts) if all_amounts else 0
    avg_invoice = round(sum(all_amounts) / len(all_amounts), 2) if all_amounts else 0

    # Supplier aggregation
    suppliers_map: dict = defaultdict(lambda: {"net": 0.0, "vat": 0.0, "count": 0})
    for r in incoming:
        k = r["issuer"] or r["receiver"] or "غير محدد"
        suppliers_map[k]["net"]   += r["net_after"] or r["net"]
        suppliers_map[k]["vat"]   += r["vat"]
        suppliers_map[k]["count"] += 1
    suppliers = sorted(
        [{"name": k, **v} for k, v in suppliers_map.items()],
        key=lambda x: x["net"], reverse=True
    )[:10]

    summary = {
        "net_vat":        net_vat,
        "sales_net":      sales_net,
        "sales_vat":      sales_vat,
        "pur_net":        pur_net,
        "pur_vat":        pur_vat,
        "incoming_count": len(incoming),
        "outgoing_count": len(outgoing),
        "total_invoices": len(rows),
        "supplier_count": len(suppliers_map),
        "max_invoice":    max_invoice,
        "avg_invoice":    avg_invoice,
        "suppliers":      suppliers,
    }

    return {
        "company_name": company_name,
        "tax_number":   tax_number,
        "period_label": period_label,
        "year":         year,
        "month":        month,
        "summary":      summary,
        "invoices":     rows,
    }


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.post("/analyze")
async def analyze(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    """Upload ETA Excel, parse it, store result, return full analysis."""
    _ensure_table(db)

    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "يُقبل ملف Excel فقط (.xlsx أو .xls)")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "حجم الملف يتجاوز الحد المسموح (20 MB)")

    try:
        parsed = _parse_excel(content)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[vat_excel] parse error: {e}", exc_info=True)
        raise HTTPException(400, f"تعذر تحليل الملف: {e}")

    s = parsed["summary"]
    # Store in DB
    row = db.execute(text("""
        INSERT INTO vat_excel_analyses
            (user_id, company_name, tax_number, period_label, year, month,
             net_vat, sales_net, sales_vat, pur_net, pur_vat, total_invoices, data_json)
        VALUES
            (:uid, :cn, :tn, :pl, :y, :m, :nv, :sn, :sv, :pn, :pv, :ti, :dj)
        RETURNING id, created_at
    """), {
        "uid": cu.id,
        "cn":  parsed["company_name"][:300] if parsed["company_name"] else "",
        "tn":  parsed["tax_number"][:80] if parsed["tax_number"] else "",
        "pl":  parsed["period_label"][:80],
        "y":   parsed["year"],
        "m":   parsed["month"],
        "nv":  s["net_vat"],
        "sn":  s["sales_net"],
        "sv":  s["sales_vat"],
        "pn":  s["pur_net"],
        "pv":  s["pur_vat"],
        "ti":  s["total_invoices"],
        "dj":  json.dumps(parsed, ensure_ascii=False, default=str),
    }).fetchone()
    db.commit()

    return {**parsed, "id": row[0], "created_at": str(row[1])}


@router.get("/history")
def history(
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    _ensure_table(db)
    rows = db.execute(text("""
        SELECT id, company_name, tax_number, period_label, year, month,
               net_vat, total_invoices, created_at
        FROM vat_excel_analyses
        WHERE user_id = :uid
        ORDER BY created_at DESC
        LIMIT 50
    """), {"uid": cu.id}).fetchall()
    return [
        {
            "id":             r[0],
            "company_name":   r[1],
            "tax_number":     r[2],
            "period_label":   r[3],
            "year":           r[4],
            "month":          r[5],
            "net_vat":        float(r[6] or 0),
            "total_invoices": r[7],
            "created_at":     str(r[8]),
        }
        for r in rows
    ]


@router.get("/history/{analysis_id}")
def get_history(
    analysis_id: int,
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    _ensure_table(db)
    row = db.execute(text("""
        SELECT id, data_json, created_at
        FROM vat_excel_analyses
        WHERE id = :id AND user_id = :uid
    """), {"id": analysis_id, "uid": cu.id}).fetchone()
    if not row:
        raise HTTPException(404, "التحليل غير موجود")
    data = json.loads(row[1])
    return {**data, "id": row[0], "created_at": str(row[2])}


@router.delete("/history/{analysis_id}")
def delete_history(
    analysis_id: int,
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    _ensure_table(db)
    result = db.execute(text("""
        DELETE FROM vat_excel_analyses
        WHERE id = :id AND user_id = :uid
    """), {"id": analysis_id, "uid": cu.id})
    db.commit()
    if result.rowcount == 0:
        raise HTTPException(404, "التحليل غير موجود")
    return {"ok": True}


def _get_analysis_data(analysis_id: int, user_id: int, db: Session) -> dict:
    _ensure_table(db)
    row = db.execute(text("""
        SELECT data_json FROM vat_excel_analyses
        WHERE id = :id AND user_id = :uid
    """), {"id": analysis_id, "uid": user_id}).fetchone()
    if not row:
        raise HTTPException(404, "التحليل غير موجود")
    return json.loads(row[0])


@router.get("/drill/{analysis_id}/{key}")
def drill(
    analysis_id: int,
    key: str,
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    data = _get_analysis_data(analysis_id, cu.id, db)
    invoices = data.get("invoices", [])

    if key == "all":
        filtered = invoices
    elif key in ("sales_net", "sales_vat"):
        filtered = [i for i in invoices if "صادرة" in i.get("direction", "")]
    elif key in ("pur_net", "pur_vat"):
        filtered = [i for i in invoices if "واردة" in i.get("direction", "")]
    else:
        filtered = invoices

    total_net = round(sum(i.get("net_after") or i.get("net") or 0 for i in filtered), 2)
    total_vat = round(sum(i.get("vat") or 0 for i in filtered), 2)
    total_amt = round(sum(i.get("total") or 0 for i in filtered), 2)

    return {
        "count":     len(filtered),
        "total_net": total_net,
        "total_vat": total_vat,
        "total_amt": total_amt,
        "invoices":  filtered,
    }


@router.get("/drill/{analysis_id}/{key}/csv")
def drill_csv(
    analysis_id: int,
    key: str,
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    data = _get_analysis_data(analysis_id, cu.id, db)
    invoices = data.get("invoices", [])

    if key in ("sales_net", "sales_vat"):
        filtered = [i for i in invoices if "صادرة" in i.get("direction", "")]
    elif key in ("pur_net", "pur_vat"):
        filtered = [i for i in invoices if "واردة" in i.get("direction", "")]
    else:
        filtered = invoices

    cols = ["uuid", "internal_id", "doc_type", "direction", "status",
            "issue_date", "issuer", "receiver", "net_after", "vat", "total", "currency"]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(filtered)

    return StreamingResponse(
        io.BytesIO(("﻿" + buf.getvalue()).encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="drill_{key}.csv"'},
    )


@router.get("/declaration/{analysis_id}")
def declaration(
    analysis_id: int,
    overrides: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    cu: User = Depends(get_current_user),
):
    """Generate VAT return declaration Excel (نموذج 10 ض.ق.م.)."""
    data    = _get_analysis_data(analysis_id, cu.id, db)
    s       = data.get("summary", {})
    ov      = {}
    if overrides:
        try:
            ov = json.loads(overrides)
        except Exception:
            pass

    sales_net = float(ov.get("sales_net", s.get("sales_net", 0)))
    sales_vat = float(ov.get("sales_vat", s.get("sales_vat", 0)))
    pur_net   = float(ov.get("pur_net",   s.get("pur_net",   0)))
    pur_vat   = float(ov.get("pur_vat",   s.get("pur_vat",   0)))
    net_vat   = round(sales_vat - pur_vat, 2)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "إقرار ض.ق.م"
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill("solid", fgColor="0F6E56")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        label_font  = Font(bold=True, size=11)
        val_font    = Font(size=11)
        border      = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right  = Alignment(horizontal="right",  vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")

        def cell(row, col, value, bold=False, fill=None, font=None, align=None):
            c = ws.cell(row=row, column=col, value=value)
            c.border  = border
            c.font    = font or (label_font if bold else val_font)
            c.alignment = align or right
            if fill:
                c.fill = fill
            return c

        # Title
        ws.merge_cells("A1:D1")
        t = ws["A1"]
        t.value     = "إقرار ضريبة القيمة المضافة (نموذج 10)"
        t.font      = Font(bold=True, size=14, color="0F6E56")
        t.alignment = center
        ws.row_dimensions[1].height = 30

        # Company info
        cell(2, 1, "اسم الشركة / الممول", bold=True)
        cell(2, 2, data.get("company_name") or "")
        cell(2, 3, "الرقم الضريبي", bold=True)
        cell(2, 4, data.get("tax_number") or "")

        cell(3, 1, "فترة الإقرار", bold=True)
        cell(3, 2, data.get("period_label") or "")
        cell(3, 3, "تاريخ الإعداد", bold=True)
        cell(3, 4, datetime.now().strftime("%Y-%m-%d"))

        ws.row_dimensions[4].height = 8

        # Table header
        for ci, label in enumerate(["البيان", "القيمة (ج.م)"], start=1):
            c = cell(5, ci, label, bold=True, fill=header_fill, font=header_font, align=center)

        # Section A
        cell(6, 1, "القسم الأول: المبيعات (ضريبة المخرجات)", bold=True)
        cell(6, 2, "")
        ws.merge_cells("A6:B6")

        rows_data = [
            ("إجمالي المبيعات الخاضعة للضريبة (بدون ضريبة)", sales_net),
            ("ضريبة المخرجات المستحقة", sales_vat),
        ]
        for ri, (lbl, val) in enumerate(rows_data, start=7):
            cell(ri, 1, lbl)
            cell(ri, 2, round(val, 2), align=left)

        r = 7 + len(rows_data) + 1  # blank row
        ws.row_dimensions[r - 1].height = 8

        # Section B
        cell(r, 1, "القسم الثاني: المشتريات (ضريبة المدخلات)", bold=True)
        cell(r, 2, "")
        ws.merge_cells(f"A{r}:B{r}")

        rows_b = [
            ("إجمالي المشتريات الخاضعة للضريبة (بدون ضريبة)", pur_net),
            ("ضريبة المدخلات المستحقة الخصم", pur_vat),
        ]
        for ri2, (lbl, val) in enumerate(rows_b, start=r + 1):
            cell(ri2, 1, lbl)
            cell(ri2, 2, round(val, 2), align=left)

        r2 = r + len(rows_b) + 2

        # Result
        result_fill = PatternFill("solid", fgColor="F0FDF4" if net_vat <= 0 else "FFF5F5")
        result_label = "رصيد دائن لصالح الممول" if net_vat <= 0 else "صافي الضريبة المستحقة للسداد"
        cell(r2, 1, result_label, bold=True)
        cell(r2, 2, round(abs(net_vat), 2), bold=True, align=left)
        ws["A" + str(r2)].fill = result_fill
        ws["B" + str(r2)].fill = result_fill

        # Column widths
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 24

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_period = (data.get("period_label") or "declaration").replace(" ", "_").replace("/", "_")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="vat_declaration_{safe_period}.xlsx"'},
        )

    except Exception as e:
        logger.error(f"[vat_excel] declaration error: {e}", exc_info=True)
        raise HTTPException(500, f"خطأ في إنشاء ملف الإقرار: {e}")
