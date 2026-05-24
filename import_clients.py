#!/usr/bin/env python3
"""
سكريبت استيراد بيانات العملاء من ملف Excel
Import clients from Excel file into the accounting system
"""

import openpyxl
import requests
import json
import time
import sys
from typing import Optional

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Configuration
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
API_BASE = "https://ms-accounting-api-production.up.railway.app"
EXCEL_FILE = "/Users/render/Downloads/بيانات تسجيل الشركات.xlsx"

# Login credentials (admin user)
LOGIN_EMAIL = "admin@ms-accounting.com"
LOGIN_PASS = "admin123"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Entity type mapping (Arabic → system code)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ENTITY_MAP = {
    "شخص واحد": "one_person",
    "مسؤلية محدودة": "llc",
    "مسؤولية محدودة": "llc",
    "ذات مسئولية محدوده": "llc",
    "ذات مسئولية محدودة": "llc",
    " ذات مسئولية محدوده": "llc",
    "مسئولية محدودة": "llc",
    "فردي": "sole",
    "منشأة فردية": "sole",
    "توصية بسيطة": "limited_partnership",
    "مساهمة": "joint_stock",
    "تضامن": "partnership",
    "فرع شركة أجنبية": "foreign_branch",
    "جمعية": "association",
    "مؤسسة": "foundation",
    "شركة قابضة": "holding",
    "منطقة حرة": "free_zone",
    "فرد": "individual",
    "شخص طبيعي": "individual",
    "عمل حر": "freelancer",
    "تحت التاسيس": "llc",  # under formation → LLC default
    "": "llc",
}


def map_entity(arabic: Optional[str]) -> str:
    if not arabic:
        return "llc"
    arabic = str(arabic).strip()
    return ENTITY_MAP.get(arabic, "llc")


def clean_fee(val) -> Optional[float]:
    """Extract monthly fee as float, return None if not a valid number."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("؟؟", "?", "", "Out", "لايوجد"):
        return None
    # Try to parse if it starts with digits
    import re
    m = re.match(r"[\d,]+", s)
    if m:
        try:
            return float(m.group().replace(",", ""))
        except:
            pass
    return None


def clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None", "لايوجد"):
        return None
    return s


def clean_tax_number(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("0", "0.0", "", "None"):
        return None
    # Remove .0 suffix from Excel numeric values
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Step 1: Login
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def login():
    print("🔐 تسجيل الدخول...")
    resp = requests.post(f"{API_BASE}/api/auth/login", data={
        "username": LOGIN_EMAIL,
        "password": LOGIN_PASS
    })
    if resp.status_code != 200:
        print(f"❌ فشل تسجيل الدخول: {resp.status_code} {resp.text}")
        sys.exit(1)
    token = resp.json().get("access_token")
    print(f"✅ تم تسجيل الدخول بنجاح")
    return token


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Step 2: Parse Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def parse_excel():
    print(f"📊 قراءة ملف Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # companies dict: name -> company data
    companies = {}

    # ─── Sheet 1: شركات القيمة المضافة ───
    # Cols: 0=م, 1=اسم العميل, 2=الكيان القانونى, 3=استمارة2, 4=قيمة مضافة,
    #       5=خصم واضافة, 6=كسب عمل, 7=القيمة المضافة (status), 8=الاتعاب الشهرية
    ws = wb['شركات القيمة المضافة']
    for row in list(ws.iter_rows(values_only=True))[2:]:
        name = clean_str(row[1])
        if not name:
            continue
        companies[name] = {
            "name": name,
            "entity": map_entity(row[2]),
            "fee": clean_fee(row[8]),
            "vat_registered": True,
            "email": None, "phone": None,
            "national_id": None, "owner_name": None,
            "commercial_register": None, "tax_number": None,
            "notes": clean_str(row[7]),  # VAT registration date
        }

    print(f"  شركات القيمة المضافة: {len(companies)} شركة")
    count_sheet1 = len(companies)

    # ─── Sheet 2: شركات خاضعه القيمة المضافة ───
    # Cols: 0=م, 1=محاسب, 2=اسم العميل, 3-14=months 1-12,
    #       15=اسم العميل(owner), 16=رقم البطاقة, 17=الكيان, 18=username, 19=password,
    #       20=email, 21=email_pass, 22=payroll_user, 23=payroll_pass,
    #       24=phone, 25=استمارة2, 26=قيمة مضافة, 27=خصم واضافة, 28=كسب عمل,
    #       29=القيمة المضافة (status), 30=الاتعاب الشهرية
    ws = wb['شركات خاضعه القيمة المضافة']
    for row in list(ws.iter_rows(values_only=True))[2:]:
        name = clean_str(row[2])
        if not name:
            continue
        email = clean_str(row[20])
        phone = clean_str(str(row[24])) if row[24] else None
        nat_id = clean_str(str(row[16])) if row[16] else None
        if nat_id and nat_id.endswith(".0"):
            nat_id = nat_id[:-2]
        owner = clean_str(row[15])
        fee = clean_fee(row[30])
        entity = map_entity(row[17])
        notes = clean_str(row[29])

        if name not in companies:
            companies[name] = {
                "name": name, "entity": entity, "fee": fee,
                "vat_registered": True,
                "email": email, "phone": phone,
                "national_id": nat_id, "owner_name": owner,
                "commercial_register": None, "tax_number": None,
                "notes": notes,
            }
        else:
            c = companies[name]
            if email and not c["email"]: c["email"] = email
            if phone and not c["phone"]: c["phone"] = phone
            if nat_id and not c["national_id"]: c["national_id"] = nat_id
            if owner and not c["owner_name"]: c["owner_name"] = owner
            if fee and not c["fee"]: c["fee"] = fee

    print(f"  بعد شركات خاضعه: {len(companies)} شركة (+{len(companies)-count_sheet1})")
    count_sheet2 = len(companies)

    # ─── Sheet 3: شركات غير خاضعه القيمة المضافة ───
    # Cols: 0=م, 1=محاسب, 2=اسم العميل, 3-14=months,
    #       15=اسم العميل(owner), 16=الكيان, 17=username, 18=password,
    #       19=email, 20=email_pass, 21=رقم البطاقة, 22=payroll_user,
    #       23=payroll_pass, 24=البيان, 25=رقم تلفون,
    #       26=استمارة2, 27=قيمة مضافة, 28=خصم واضافة, 29=كسب عمل,
    #       30=القيمة المضافة (status), 31=الاتعاب الشهرية
    ws = wb['شركات غير خاضعه القيمة المضافة']
    for row in list(ws.iter_rows(values_only=True))[2:]:
        name = clean_str(row[2])
        if not name:
            continue
        entity = map_entity(row[16])
        fee = clean_fee(row[32]) if len(row) > 32 else None
        email = clean_str(row[19]) if len(row) > 19 else None
        phone = clean_str(str(row[25])) if len(row) > 25 and row[25] else None
        owner = clean_str(row[15]) if len(row) > 15 else None
        nat_id = clean_str(str(row[21])) if len(row) > 21 and row[21] else None
        if nat_id and nat_id.endswith(".0"):
            nat_id = nat_id[:-2]
        notes = clean_str(row[30]) if len(row) > 30 else None

        if name not in companies:
            companies[name] = {
                "name": name, "entity": entity, "fee": fee,
                "vat_registered": False,
                "email": email, "phone": phone,
                "national_id": nat_id, "owner_name": owner,
                "commercial_register": None, "tax_number": None,
                "notes": notes,
            }
        else:
            c = companies[name]
            if email and not c["email"]: c["email"] = email
            if phone and not c["phone"]: c["phone"] = phone
            if nat_id and not c["national_id"]: c["national_id"] = nat_id

    print(f"  بعد شركات غير خاضعه: {len(companies)} شركة (+{len(companies)-count_sheet2})")
    count_sheet3 = len(companies)

    # ─── Sheet 4: بوابة قديمة (enrich with commercial_register + tax_number) ───
    # Cols: 0=م, 1=اسم الشركة, 2=اسم الممول, 3=رقم البطاقة,
    #       4=الكيان, 5=رقم السجل التجاري, 6=رقم تاميني,
    #       7=رقم التسجيل الضريبي, 8=اسم المستخدم, 9=باسورد,
    #       10=ايميل
    ws = wb['بوابة قديمة']
    portal_companies = {}
    for row in list(ws.iter_rows(values_only=True))[2:]:
        pname = clean_str(row[1])
        if not pname:
            continue
        portal_companies[pname] = {
            "owner": clean_str(row[2]),
            "national_id": clean_str(str(row[3])) if row[3] else None,
            "entity": map_entity(row[4]),
            "commercial_register": clean_str(str(row[5])) if row[5] else None,
            "tax_number": clean_tax_number(row[7]),
            "email": clean_str(row[10]),
        }
        # Fix national_id
        if portal_companies[pname]["national_id"] and portal_companies[pname]["national_id"].endswith(".0"):
            portal_companies[pname]["national_id"] = portal_companies[pname]["national_id"][:-2]

    # Merge portal data into main companies
    merged = 0
    for pname, pdata in portal_companies.items():
        best_match = None
        best_score = 0
        for cname in companies:
            # Score based on name overlap
            score = 0
            if pname == cname:
                score = 100
            elif pname in cname or cname in pname:
                score = 80
            elif pname[:6] == cname[:6]:
                score = 60
            elif pname[:4] == cname[:4]:
                score = 40
            if score > best_score:
                best_score = score
                best_match = cname

        if best_match and best_score >= 60:
            c = companies[best_match]
            if pdata["commercial_register"] and not c["commercial_register"]:
                c["commercial_register"] = pdata["commercial_register"]
            if pdata["tax_number"] and not c["tax_number"]:
                c["tax_number"] = pdata["tax_number"]
            if pdata["owner"] and not c["owner_name"]:
                c["owner_name"] = pdata["owner"]
            if pdata["email"] and not c["email"]:
                c["email"] = pdata["email"]
            if pdata["national_id"] and not c["national_id"]:
                c["national_id"] = pdata["national_id"]
            merged += 1
        else:
            # Add portal company as new client (they're existing clients)
            if pname not in companies:
                companies[pname] = {
                    "name": pname,
                    "entity": pdata["entity"],
                    "fee": None,
                    "vat_registered": False,
                    "email": pdata["email"],
                    "phone": None,
                    "national_id": pdata["national_id"],
                    "owner_name": pdata["owner"],
                    "commercial_register": pdata["commercial_register"],
                    "tax_number": pdata["tax_number"],
                    "notes": None,
                }

    print(f"  بعد إثراء بيانات البوابة القديمة: {len(companies)} شركة")
    print(f"  تم دمج بيانات {merged} شركة من البوابة القديمة")

    return list(companies.values())


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Step 3: Get existing clients to avoid duplicates
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def get_existing_clients(token: str) -> dict:
    headers = {"Authorization": f"Bearer {token}"}
    print("📋 جلب قائمة العملاء الحاليين...")
    existing = {}
    page = 1
    while True:
        resp = requests.get(f"{API_BASE}/api/clients",
                          params={"page": page, "page_size": 100},
                          headers=headers)
        if resp.status_code != 200:
            break
        data = resp.json()
        items = data.get("items", data) if isinstance(data, dict) else data
        if not items:
            break
        for c in items:
            existing[c["name"].strip()] = c["id"]
        if isinstance(data, dict) and len(items) < 100:
            break
        page += 1
    print(f"  عملاء موجودون: {len(existing)}")
    return existing


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Step 4: Create clients
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def determine_obligations(company: dict) -> list:
    """Determine tax obligations based on company type and VAT registration."""
    obligations = []

    if company.get("vat_registered"):
        # VAT registered companies
        obligations.extend(["vat_monthly", "withholding_monthly"])

    # All companies with employees might have:
    # - payroll_monthly (if they have employees)
    # - work_profit
    # - social_insurance

    # Income tax for all
    obligations.append("income_quarterly")
    obligations.append("annual_declaration")

    return obligations


def create_clients(companies: list, existing: dict, token: str) -> dict:
    """Create clients via API. Returns dict of name -> id."""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    client_ids = dict(existing)  # copy existing

    created = 0
    skipped = 0
    errors = 0

    print(f"\n🚀 بدء إنشاء {len(companies)} عميل...")

    for i, company in enumerate(companies, 1):
        name = company["name"]

        # Skip if already exists
        if name in existing:
            skipped += 1
            continue

        # Determine tax type
        tax_type = "vat" if company.get("vat_registered") else "income"

        # Determine obligations
        obligations = determine_obligations(company)

        # Build notes
        notes_parts = []
        if company.get("notes"):
            notes_parts.append(company["notes"])
        if company.get("owner_name"):
            notes_parts.append(f"اسم الممول: {company['owner_name']}")

        payload = {
            "name": name,
            "client_type": company["entity"],
            "status": "active",
            "email": company.get("email"),
            "phone": company.get("phone"),
            "national_id": company.get("national_id"),
            "commercial_register": company.get("commercial_register"),
            "tax_number": company.get("tax_number"),
            "tax_type": tax_type,
            "monthly_fee": company.get("fee") or 0,
            "tax_obligations": obligations,
            "notes": "\n".join(notes_parts) if notes_parts else None,
        }

        # Remove None values
        payload = {k: v for k, v in payload.items() if v is not None}

        resp = requests.post(f"{API_BASE}/api/clients",
                           json=payload, headers=headers)

        if resp.status_code == 200:
            client_id = resp.json().get("id")
            client_ids[name] = client_id
            created += 1
            print(f"  ✅ [{i}/{len(companies)}] {name} (ID: {client_id})")
        else:
            errors += 1
            print(f"  ❌ [{i}/{len(companies)}] {name}: {resp.status_code} {resp.text[:100]}")

        # Small delay to avoid overwhelming the API
        time.sleep(0.3)

    print(f"\n📊 النتيجة:")
    print(f"  تم إنشاء: {created} عميل")
    print(f"  موجود مسبقاً: {skipped} عميل")
    print(f"  أخطاء: {errors} عميل")

    return client_ids


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Step 5: Create collection contracts for monthly fee clients
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_collection_contracts(companies: list, client_ids: dict, token: str):
    """Create monthly fee collection contracts for companies with fees."""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    fee_companies = [c for c in companies if c.get("fee") and c["fee"] > 0]
    print(f"\n💰 إنشاء عقود التحصيل الشهري لـ {len(fee_companies)} شركة...")

    # Get existing contracts
    resp = requests.get(f"{API_BASE}/api/collections", headers=headers)
    existing_contracts = set()
    if resp.status_code == 200:
        data = resp.json()
        items = data.get("items", data) if isinstance(data, dict) else data
        for c in items:
            existing_contracts.add(c.get("client_id"))

    created = 0
    skipped = 0
    errors = 0

    for company in fee_companies:
        name = company["name"]
        client_id = client_ids.get(name)

        if not client_id:
            print(f"  ⚠️  لا يوجد ID للعميل: {name}")
            continue

        if client_id in existing_contracts:
            skipped += 1
            continue

        payload = {
            "client_id": client_id,
            "collection_type": "monthly_fee",
            "title": f"أتعاب محاسبة شهرية - {name}",
            "agreed_amount": company["fee"] * 12,  # Annual amount
            "monthly_amount": company["fee"],
            "is_recurring": True,
            "recurring_day": 1,
            "service_description": "خدمات محاسبة شهرية شاملة",
            "is_active": True,
        }

        resp = requests.post(f"{API_BASE}/api/collections",
                           json=payload, headers=headers)

        if resp.status_code == 200:
            created += 1
            print(f"  ✅ عقد: {name} ({company['fee']} ج.م/شهر)")
        else:
            errors += 1
            print(f"  ❌ فشل عقد {name}: {resp.status_code} {resp.text[:100]}")

        time.sleep(0.2)

    print(f"\n📊 عقود التحصيل:")
    print(f"  تم إنشاء: {created}")
    print(f"  موجود مسبقاً: {skipped}")
    print(f"  أخطاء: {errors}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Main
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print("=" * 60)
    print("  استيراد بيانات العملاء من ملف Excel")
    print("=" * 60)
    print()

    # Login
    token = login()

    # Parse Excel
    companies = parse_excel()
    print(f"\n✅ إجمالي الشركات المستخلصة: {len(companies)}")

    # Show summary
    with_fee = [c for c in companies if c.get("fee") and c["fee"] > 0]
    vat_cos = [c for c in companies if c.get("vat_registered")]
    print(f"  منها مسجلة ضريبة القيمة المضافة: {len(vat_cos)}")
    print(f"  منها لها أتعاب شهرية: {len(with_fee)}")

    # Get existing clients
    existing = get_existing_clients(token)

    # Create clients
    client_ids = create_clients(companies, existing, token)

    # Create collection contracts
    create_collection_contracts(companies, client_ids, token)

    print("\n" + "=" * 60)
    print("✅ تم الاستيراد بنجاح!")
    print("=" * 60)


if __name__ == "__main__":
    main()
